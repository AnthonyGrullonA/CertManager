"""Generación y exportación de reportes de certificados (síncrono).

Este módulo contiene la lógica de dominio del constructor de reportes,
desacoplada de las vistas web:

- ``ReportFilters``: normaliza los filtros del formulario (rango de fechas,
  grupos, estados, ventana de vencimiento, emisor, responsable).
- ``build_report``: aplica los filtros sobre los certificados visibles del
  usuario (scoping) y devuelve un ``ReportResult`` con el queryset, la
  distribución por estado (donut), las ventanas de vencimiento (barras) y los
  totales — lo que la preview en vivo renderiza.
- ``export_*``: serializa el ``ReportResult`` a PDF (reportlab), Excel
  (openpyxl) y CSV (stdlib). Multi-formato simultáneo se obtiene llamando a
  ``build_export`` con varios formatos (se empaqueta en ZIP).

El envío real de los reportes programados se difiere a un management-command
(Celery DIFERIDO): aquí solo vive la generación y exportación síncronas.
"""
from __future__ import annotations

import csv
import io
import zipfile
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta

from django.utils import timezone

from apps.certificates.models import Certificate
from apps.core.enums import CertificateStatus, ReportFormat

# Ventanas de vencimiento (días) para el gráfico de barras de la preview.
EXPIRY_WINDOWS = [7, 15, 30, 60, 90]

# Paleta Forge UI por estado (hex), espejo del dashboard. Sirve para el PDF.
STATE_COLORS = {
    "VIGENTE": "#10b981",
    "POR_VENCER": "#f59e0b",
    "CRITICO": "#f97316",
    "VENCIDO": "#ef4444",
    "ERROR": "#8b5cf6",
    "SIN_CHEQUEAR": "#94a3b8",
}

# Rangos de fecha predefinidos (clave -> días hacia atrás). None = sin límite.
DATE_RANGES = {
    "7": 7,
    "30": 30,
    "90": 90,
    "365": 365,
    "all": None,
}
DATE_RANGE_LABELS = {
    "7": "Últimos 7 días",
    "30": "Últimos 30 días",
    "90": "Últimos 90 días",
    "365": "Último año",
    "all": "Todo el historial",
}


def _parse_date(value):
    """Convierte una cadena ISO (YYYY-MM-DD) a date, o None si no aplica."""
    if not value:
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    try:
        return datetime.strptime(str(value).strip(), "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None


@dataclass
class ReportFilters:
    """Filtros normalizados del constructor de reportes."""

    template: str = "INVENTORY"
    date_range: str = "30"
    date_from: date | None = None
    date_to: date | None = None
    teams: list[int] = field(default_factory=list)
    statuses: list[str] = field(default_factory=list)
    expiry_window: str = ""          # "", "7", "15", "30", "90"
    issuer: str = ""
    recipient: str = ""              # correo del responsable/destinatario

    @classmethod
    def from_request(cls, data, *, multi=True):
        """Construye filtros desde un QueryDict (GET/POST).

        ``multi`` controla si grupos/estados se leen como listas (getlist).
        """
        get_list = data.getlist if hasattr(data, "getlist") else (lambda k: data.get(k, []))

        teams = []
        for raw in (get_list("teams") if multi else [data.get("teams")]):
            try:
                if raw not in (None, ""):
                    teams.append(int(raw))
            except (TypeError, ValueError):
                continue

        statuses = [
            s for s in (get_list("statuses") if multi else [data.get("statuses")])
            if s in CertificateStatus.values
        ]

        date_range = (data.get("date_range") or "30").strip()
        if date_range not in DATE_RANGES:
            date_range = "30"

        return cls(
            template=(data.get("template") or "INVENTORY").strip().upper(),
            date_range=date_range,
            date_from=_parse_date(data.get("date_from")),
            date_to=_parse_date(data.get("date_to")),
            teams=teams,
            statuses=statuses,
            expiry_window=(data.get("expiry_window") or "").strip(),
            issuer=(data.get("issuer") or "").strip(),
            recipient=(data.get("recipient") or "").strip(),
        )

    @property
    def date_range_label(self):
        if self.date_from or self.date_to:
            desde = self.date_from.isoformat() if self.date_from else "inicio"
            hasta = self.date_to.isoformat() if self.date_to else "hoy"
            return f"{desde} → {hasta}"
        return DATE_RANGE_LABELS.get(self.date_range, "Últimos 30 días")


@dataclass
class ReportResult:
    """Resultado de un reporte: certificados + agregados para la preview."""

    filters: ReportFilters
    certificates: list
    total: int
    status_distribution: list   # [{key,label,value}]
    expiry_windows: list        # [{label,value}]
    template_label: str
    scope_label: str = "Todos los grupos"
    # Tabla de detalle del reporte: columnas + filas dependen de la PLANTILLA
    # (inventario/por-vencer/vencidos = certs; historial = chequeos; por-grupo =
    # agregado). El total_label describe qué cuenta `total` (certs / chequeos /
    # grupos) para que la preview y los exportes muestren datos certeros.
    columns: list = field(default_factory=list)   # [str]
    rows: list = field(default_factory=list)       # [[str|num, …]]
    total_label: str = "certificados"

    @property
    def kpis(self):
        counts = {d["key"]: d["value"] for d in self.status_distribution}
        return {
            "total": self.total,
            "vigente": counts.get("VIGENTE", 0),
            "por_vencer": counts.get("POR_VENCER", 0),
            "critico": counts.get("CRITICO", 0),
            "vencido": counts.get("VENCIDO", 0),
            "error": counts.get("ERROR", 0),
            "sin_chequear": counts.get("SIN_CHEQUEAR", 0),
        }


def _apply_filters(qs, filters: ReportFilters):
    """Aplica los filtros del constructor sobre un queryset de certificados."""
    # Plantilla: presets de estado que se suman a los filtros explícitos.
    if filters.template == "EXPIRING":
        qs = qs.filter(status__in=[CertificateStatus.POR_VENCER, CertificateStatus.CRITICO])
    elif filters.template == "EXPIRED":
        qs = qs.filter(status__in=[CertificateStatus.VENCIDO, CertificateStatus.ERROR])

    # Grupos.
    if filters.teams:
        qs = qs.filter(team_id__in=filters.teams)

    # Estados explícitos.
    if filters.statuses:
        qs = qs.filter(status__in=filters.statuses)

    # Ventana de vencimiento (días restantes <= N, incluye vencidos).
    if filters.expiry_window:
        try:
            window = int(filters.expiry_window)
            qs = qs.filter(days_left__isnull=False, days_left__lte=window)
        except (TypeError, ValueError):
            pass

    # Emisor (coincidencia parcial).
    if filters.issuer:
        qs = qs.filter(issuer__icontains=filters.issuer)

    # Responsable / destinatario (correo del recipient).
    if filters.recipient:
        qs = qs.filter(recipients__email__icontains=filters.recipient)

    # Rango de fechas: por defecto sobre last_checked_at; si no hay chequeo,
    # caemos a created_at para no perder certificados sin chequear.
    desde = filters.date_from
    hasta = filters.date_to
    if not desde and not hasta:
        days = DATE_RANGES.get(filters.date_range)
        if days is not None:
            desde = (timezone.now() - timedelta(days=days)).date()
    if desde:
        qs = qs.filter(created_at__date__gte=desde)
    if hasta:
        qs = qs.filter(created_at__date__lte=hasta)

    return qs.distinct()


def build_report(user, filters: ReportFilters, *, scope_label="Todos los grupos") -> ReportResult:
    """Genera el reporte aplicando scoping del usuario + filtros del constructor."""
    qs = (
        Certificate.objects.for_user(user)
        .select_related("team")
        .order_by("days_left")
    )
    qs = _apply_filters(qs, filters)

    certs = list(qs)

    # Distribución por estado (donut).
    counts = {}
    for c in certs:
        counts[c.status] = counts.get(c.status, 0) + 1
    status_distribution = [
        {"key": s.value, "label": str(s.label), "value": counts.get(s.value, 0)}
        for s in CertificateStatus
    ]

    # Ventanas de vencimiento (barras, buckets no solapados, ≤7d incluye vencidos).
    windows = []
    prev = None
    for w in EXPIRY_WINDOWS:
        n = 0
        for c in certs:
            d = c.days_left
            if d is None:
                continue
            if prev is None:
                if d < w:
                    n += 1
            elif prev <= d < w:
                n += 1
        windows.append({"label": f"≤{w}d", "value": n})
        prev = w

    from apps.core.enums import ReportTemplate
    template_label = str(dict(ReportTemplate.choices).get(filters.template, "Inventario de certificados"))

    # Tabla de detalle según la PLANTILLA: cada una produce columnas/filas
    # distintas (datos certeros, no siempre el mismo reporte).
    columns, rows, total, total_label = _build_table(filters, certs, user)

    return ReportResult(
        filters=filters,
        certificates=certs,
        total=total,
        status_distribution=status_distribution,
        expiry_windows=windows,
        template_label=template_label,
        scope_label=scope_label,
        columns=columns,
        rows=rows,
        total_label=total_label,
    )


# Encabezados de la tabla de certificados (inventario / por-vencer / vencidos).
_CERT_COLS = ["Dominio", "Puerto", "Grupo", "Estado", "Días restantes", "Vence", "Emisor", "Último chequeo"]


def _cert_row(cert):
    return [
        cert.domain,
        str(cert.port),
        cert.team.name if cert.team_id else "",
        str(cert.get_status_display()),
        "" if cert.days_left is None else str(cert.days_left),
        cert.valid_to.strftime("%Y-%m-%d") if cert.valid_to else "",
        cert.issuer or "",
        cert.last_checked_at.strftime("%Y-%m-%d %H:%M") if cert.last_checked_at else "Sin chequear",
    ]


def _build_table(filters, certs, user):
    """Devuelve (columns, rows, total, total_label) según la plantilla."""
    tpl = filters.template

    if tpl == "HISTORY":
        # Historial: últimos chequeos de los certificados del ámbito.
        from apps.certificates.models import CertificateCheck

        cert_ids = [c.id for c in certs]
        checks = (
            CertificateCheck.objects.filter(certificate_id__in=cert_ids)
            .select_related("certificate")
            .order_by("-checked_at")[:500]
        )
        cols = ["Fecha", "Dominio", "Estado", "Días", "Latencia (ms)", "Error"]
        rows = [
            [
                ch.checked_at.strftime("%Y-%m-%d %H:%M"),
                ch.certificate.domain,
                str(ch.get_status_display()),
                "" if ch.days_left is None else str(ch.days_left),
                "" if ch.latency_ms is None else str(ch.latency_ms),
                (ch.error_message or "")[:80],
            ]
            for ch in checks
        ]
        return cols, rows, len(rows), "chequeos"

    if tpl == "BY_GROUP":
        # Resumen agregado por grupo.
        from collections import defaultdict

        agg = defaultdict(lambda: {s.value: 0 for s in CertificateStatus})
        names = {}
        for c in certs:
            key = c.team_id
            names[key] = c.team.name if c.team_id else "Sin grupo"
            agg[key][c.status] += 1
        cols = ["Grupo", "Total", "Vigentes", "Por vencer", "Críticos", "Vencidos", "Errores", "Sin chequear"]
        rows = []
        for key, counts in sorted(agg.items(), key=lambda kv: names[kv[0]]):
            total = sum(counts.values())
            rows.append([
                names[key], str(total),
                str(counts["VIGENTE"]), str(counts["POR_VENCER"]), str(counts["CRITICO"]),
                str(counts["VENCIDO"]), str(counts["ERROR"]), str(counts["SIN_CHEQUEAR"]),
            ])
        return cols, rows, len(rows), "grupos"

    # INVENTORY / EXPIRING / EXPIRED: tabla de certificados (el set ya viene
    # filtrado por _apply_filters según la plantilla).
    return _CERT_COLS, [_cert_row(c) for c in certs], len(certs), "certificados"


# ---------------------------------------------------------------------------
# Exportadores (usan result.columns / result.rows según la plantilla)
# ---------------------------------------------------------------------------


def export_csv(result: ReportResult) -> bytes:
    """Serializa el reporte a CSV (UTF-8 con BOM para Excel en es-DO)."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([f"Reporte: {result.template_label}"])
    writer.writerow([f"Ámbito: {result.scope_label} · {result.filters.date_range_label} · {result.total} {result.total_label}"])
    writer.writerow([])
    writer.writerow(result.columns)
    for row in result.rows:
        writer.writerow(row)
    return ("﻿" + buf.getvalue()).encode("utf-8")


def export_excel(result: ReportResult) -> bytes:
    """Serializa el reporte a un libro Excel (.xlsx) con openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    ws.append([result.template_label])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"{result.scope_label} · {result.filters.date_range_label} · {result.total} {result.total_label}"])
    ws.append([])

    header_row = ws.max_row + 1
    ws.append(result.columns)
    bold = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="334155")
    for col in range(1, len(result.columns) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = bold
        cell.fill = fill

    for row in result.rows:
        ws.append(row)

    # Ancho de columna razonable (uniforme; se adapta a cualquier plantilla).
    for i in range(1, len(result.columns) + 1):
        ws.column_dimensions[chr(64 + i)].width = 20

    # Hoja de resumen por estado.
    ws2 = wb.create_sheet("Resumen")
    ws2.append(["Estado", "Cantidad"])
    ws2["A1"].font = Font(bold=True)
    ws2["B1"].font = Font(bold=True)
    for d in result.status_distribution:
        ws2.append([d["label"], d["value"]])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def export_pdf(result: ReportResult) -> bytes:
    """Serializa el reporte a PDF con identidad CertManager y foco operativo."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    out = io.BytesIO()
    doc = SimpleDocTemplate(
        out,
        pagesize=landscape(A4),
        topMargin=12 * mm,
        bottomMargin=14 * mm,
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        title=str(result.template_label),
        pageCompression=0,
    )
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name="CfBrand",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=18,
        leading=22,
        textColor=colors.white,
    ))
    styles.add(ParagraphStyle(
        name="CfSmallWhite",
        parent=styles["Normal"],
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#dbeafe"),
    ))
    styles.add(ParagraphStyle(
        name="CfSection",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=10,
        leading=12,
        textColor=colors.HexColor("#0f172a"),
        spaceAfter=4,
    ))
    styles.add(ParagraphStyle(
        name="CfCell",
        parent=styles["Normal"],
        fontSize=7,
        leading=9,
        textColor=colors.HexColor("#334155"),
    ))
    styles.add(ParagraphStyle(
        name="CfHeaderCell",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=7,
        leading=9,
        textColor=colors.white,
    ))

    generated_at = timezone.localtime().strftime("%Y-%m-%d %H:%M")
    subtitle = (
        "Salida de CertManager - aplicativo de monitoreo de certificados<br/>"
        f"{result.template_label} | {result.scope_label} | {result.filters.date_range_label}"
    )
    header = Table(
        [[
            Paragraph("CertManager", styles["CfBrand"]),
            Paragraph(subtitle, styles["CfSmallWhite"]),
            Paragraph(f"Generado<br/>{generated_at}", styles["CfSmallWhite"]),
        ]],
        colWidths=[42 * mm, 178 * mm, 34 * mm],
    )
    header.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#0f172a")),
        ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#0f172a")),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 9),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (2, 0), (2, 0), "RIGHT"),
    ]))
    elements = [header, Spacer(1, 6 * mm)]

    kpis = result.kpis
    kpi_data = [
        ["Total", "Vigentes", "Por vencer", "Criticos", "Vencidos", "Error / s.c."],
        [
            str(result.total),
            str(kpis["vigente"]),
            str(kpis["por_vencer"]),
            str(kpis["critico"]),
            str(kpis["vencido"]),
            str(kpis["error"] + kpis["sin_chequear"]),
        ],
    ]
    kpi_table = Table(kpi_data, colWidths=[42 * mm] * 6)
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f8fafc")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#64748b")),
        ("TEXTCOLOR", (0, 1), (-1, 1), colors.HexColor("#0f172a")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 7),
        ("FONTSIZE", (0, 1), (-1, 1), 15),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#e2e8f0")),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    elements += [kpi_table, Spacer(1, 7 * mm)]

    # Resumen por estado.
    elements.append(Paragraph("Resumen por estado", styles["CfSection"]))
    summary = [["Estado", "Cantidad"]] + [
        [d["label"], str(d["value"])] for d in result.status_distribution
    ]
    summary_table = Table(summary, colWidths=[60 * mm, 30 * mm])
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#334155")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
            ]
        )
    )
    elements += [summary_table, Spacer(1, 7 * mm)]

    # Tabla de detalle (columnas/filas según la plantilla).
    elements.append(Paragraph(f"Detalle del reporte ({result.total} {result.total_label})", styles["CfSection"]))
    data = [[Paragraph(str(c), styles["CfHeaderCell"]) for c in result.columns]]
    data += [
        [Paragraph(str(cell or "-"), styles["CfCell"]) for cell in row]
        for row in result.rows
    ]
    if len(data) == 1:
        data.append(
            [Paragraph("Sin registros que coincidan con los filtros.", styles["CfCell"])]
            + [Paragraph("", styles["CfCell"]) for _ in result.columns[1:]]
        )
    first_width = 46 * mm
    remaining = max(1, len(result.columns) - 1)
    col_widths = [first_width] + [(254 * mm - first_width) / remaining] * remaining
    detail = Table(data, repeatRows=1, colWidths=col_widths)
    detail.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#334155")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#e2e8f0")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    elements.append(detail)

    def _footer(canvas, _doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.HexColor("#64748b"))
        canvas.drawString(14 * mm, 8 * mm, "CertManager - monitoreo de certificados")
        canvas.drawRightString(283 * mm, 8 * mm, f"Pagina {canvas.getPageNumber()}")
        canvas.restoreState()

    doc.build(elements, onFirstPage=_footer, onLaterPages=_footer)
    return out.getvalue()


# Registro: formato -> (función, content_type, extensión).
EXPORTERS = {
    ReportFormat.PDF: (export_pdf, "application/pdf", "pdf"),
    ReportFormat.EXCEL: (export_excel, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx"),
    ReportFormat.CSV: (export_csv, "text/csv", "csv"),
}


def normalize_formats(formats):
    """Devuelve la lista de formatos válidos (en orden PDF, Excel, CSV)."""
    requested = {str(f).upper() for f in (formats or [])}
    return [f for f in (ReportFormat.PDF, ReportFormat.EXCEL, ReportFormat.CSV) if f in requested]


def build_export(result: ReportResult, formats):
    """Construye la descarga para uno o varios formatos.

    Devuelve ``(content_bytes, content_type, filename)``. Con un solo formato
    entrega el archivo directo; con varios los empaqueta en un ZIP (multi-formato
    simultáneo, decisión congelada nº 6 del plan).
    """
    fmts = normalize_formats(formats)
    if not fmts:
        fmts = [ReportFormat.CSV]

    slug = (result.filters.template or "reporte").lower()

    if len(fmts) == 1:
        fn, content_type, ext = EXPORTERS[fmts[0]]
        return fn(result), content_type, f"reporte-{slug}.{ext}"

    # Multi-formato: ZIP con un archivo por formato.
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for fmt in fmts:
            fn, _ct, ext = EXPORTERS[fmt]
            zf.writestr(f"reporte-{slug}.{ext}", fn(result))
    return zip_buf.getvalue(), "application/zip", f"reporte-{slug}.zip"
